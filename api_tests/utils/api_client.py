import requests
import logging
from datetime import datetime
from openpyxl import Workbook
import pytest
from openpyxl.styles import PatternFill  
import os  
from openpyxl.styles import Font, PatternFill


# API Base URL
BASE_URL = "https://petstore.swagger.io/v2"

# Set up logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger("PetstoreAPI")

class PetstoreAPI:
    """Handles API requests for Petstore endpoints"""

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({"Content-Type": "application/json"})  # Ensure JSON format for requests
        self.test_results = []  # Store test results

    def save_test_result(self, test_name, expected_status, actual_status):
        """Save test results in a structured format."""
        status = "Passed" if expected_status == actual_status else "Failed"
        self.test_results.append({
            "Test Name": test_name,
            "Expected Result": f"Status Code {expected_status}",
            "Actual Result": f"Status Code {actual_status}",
            "Status": status,
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

    def generate_excel_report(self):
        """Generate an Excel test report."""
        if not self.test_results:
            logger.warning("No test results available to generate a report.")
            return

        # Create a new workbook and get the active sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Test Results"

        # Define styles
        header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light Blue
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light Green for Passed
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light Red for Failed
        header_font = Font(bold=True)

        # Add headers to the first row  
        headers = ["Test Name", "Expected Result", "Actual Result", "Status", "Timestamp"]
        sheet.append(headers)

        # Style headers
        for cell in sheet[1]:  # First row
            cell.fill = header_fill
            cell.font = header_font

        # Add test results to the sheet
        for result in self.test_results:
            row = [
                result["Test Name"],
                result["Expected Result"],
                result["Actual Result"],
                result["Status"],
                result["Timestamp"],
            ]
            sheet.append(row)

            # Apply conditional coloring based on status
            status_cell = sheet.cell(row=sheet.max_row, column=4)  # Status column
            if result["Status"] == "Passed":
                status_cell.fill = pass_fill
            else:
                status_cell.fill = fail_fill

        # Ensure the reports directory exists
        report_dir = "api_tests/reports"
        os.makedirs(report_dir, exist_ok=True)

        # Save the workbook to the specified directory
        report_filename = os.path.join(report_dir, "Test_Report.xlsx")
        workbook.save(report_filename)
        logger.info(f"Test report generated and saved to: {report_filename}")



    def create_pet(self, pet_data):
        """Create a new pet"""
        url = f"{BASE_URL}/pet"
        response = self.session.post(url, json=pet_data)
        logger.info(f"Create Pet Response: {response.status_code} | Response Body: {response.text}")
        return response

    def get_pet(self, pet_id):
        """Retrieve pet details"""
        url = f"{BASE_URL}/pet/{pet_id}"
        response = self.session.get(url)
        logger.info(f"Get Pet Response: {response.status_code} | Response Body: {response.text}")
        return response

    def update_pet(self, pet_data):
        """Update an existing pet"""
        url = f"{BASE_URL}/pet"
        response = self.session.put(url, json=pet_data)
        logger.info(f"Update Pet Response: {response.status_code} | Response Body: {response.text}")
        return response

    def delete_pet(self, pet_id):
        """Delete a pet"""
        url = f"{BASE_URL}/pet/{pet_id}"
        response = self.session.delete(url)
        logger.info(f"Delete Pet Response: {response.status_code} | Response Body: {response.text}")
        return response

    def find_pets_by_status(self, status="available"):
        """Find pets by status (available, pending, sold)"""
        url = f"{BASE_URL}/pet/findByStatus"
        response = self.session.get(url, params={"status": status})
        logger.info(f"Find Pets by Status Response: {response.status_code} | Response Body: {response.text}")
        return response